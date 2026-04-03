import csv
import json
import sys
import threading
from argparse import ArgumentParser, Namespace
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import requests

from modules.info_reader import InfoReader
from modules.scrapper import Scrapper

BANNER: str = """
▄▄▄█████▓ ██░ ██ ▓█████   ██████  ▄████▄   ██▀███   ▄▄▄       ██▓███   ██▓███  ▓█████  ██▀███  
▓  ██▒ ▓▒▓██░ ██▒▓█   ▀ ▒██    ▒ ▒██▀ ▀█  ▓██ ▒ ██▒▒████▄    ▓██░  ██▒▓██░  ██▒▓█   ▀ ▓██ ▒ ██▒
▒ ▓██░ ▒░▒██▀▀██░▒███   ░ ▓██▄   ▒▓█    ▄ ▓██ ░▄█ ▒▒██  ▀█▄  ▓██░ ██▓▒▓██░ ██▓▒▒███   ▓██ ░▄█ ▒
░ ▓██▓ ░ ░▓█ ░██ ▒▓█  ▄   ▒   ██▒▒▓▓▄ ▄██▒▒██▀▀█▄  ░██▄▄▄▄██ ▒██▄█▓▒ ▒▒██▄█▓▒ ▒▒▓█  ▄ ▒██▀▀█▄  
  ▒██▒ ░ ░▓█▒░██▓░▒████▒▒██████▒▒▒ ▓███▀ ░░██▓ ▒██▒ ▓█   ▓██▒▒██▒ ░  ░▒██▒ ░  ░░▒████▒░██▓ ▒██▒
  ▒ ░░    ▒ ░░▒░▒░░ ▒░ ░▒ ▒▓▒ ▒ ░░ ░▒ ▒  ░░ ▒▓ ░▒▓░ ▒▒   ▓▒█░▒▓▒░ ░  ░▒▓▒░ ░  ░░░ ▒░ ░░ ▒▓ ░▒▓░
    ░     ▒ ░▒░ ░ ░ ░  ░░ ░▒  ░ ░  ░  ▒     ░▒ ░ ▒░  ▒   ▒▒ ░░▒ ░     ░▒ ░      ░ ░  ░  ░▒ ░ ▒░
  ░       ░  ░░ ░   ░   ░  ░  ░  ░          ░░   ░   ░   ▒   ░░       ░░          ░     ░░   ░ 
          ░  ░  ░   ░  ░      ░  ░ ░         ░           ░  ░                     ░  ░   ░     
                                 ░                                                            
"""


def build_parser() -> ArgumentParser:
    p = ArgumentParser(description="TheScrapper - Contact finder")
    p.add_argument("-u",   "--url",            help="Target URL.")
    p.add_argument("-us",  "--urls",           help="File containing target URLs (one per line).")
    p.add_argument("-c",   "--crawl",          action="store_true", help="Crawl every URL found on the site.")
    p.add_argument("-b",   "--banner",         action="store_true", help="Suppress the banner.")
    p.add_argument("-e",   "--email",          action="store_true", help="Extract email addresses.")
    p.add_argument("-n",   "--number",         action="store_true", help="Extract phone numbers.")
    p.add_argument("-s",   "--socials",        action="store_true", help="Extract and print social media links.")
    p.add_argument("--social-extract",         action="store_true", help="Fetch detailed info for found social media links (implies --socials).")
    p.add_argument("-o",   "--output",         action="store_true", help="Save output to a JSON file.")
    p.add_argument("--csv",                    help="CSV or Excel (.xlsx) file with URLs. Reads URLs and writes results back.")
    p.add_argument("--csv-column",             default="url", help="Column name containing URLs in the CSV/Excel file (default: 'url').")
    p.add_argument("-v",   "--verbose",        action="store_true", help="Verbose output.")
    p.add_argument("-t",   "--threads",        type=int, default=5, help="Number of concurrent threads for batch scraping (default: 5).")
    return p


def normalize_url(url: str) -> str:
    if not url.startswith(("http://", "https://")):
        return "http://" + url
    return url


def scrape(url: str, args: Namespace) -> dict:
    """Scrape a single URL and return structured results."""
    scrap = Scrapper(url=url, crawl=args.crawl)
    ir = InfoReader(content=scrap.getText())

    extract_contacts = not args.email and not args.number and not args.socials
    want_sm = args.socials or args.social_extract

    result: dict = {"Target": url}

    if args.email or extract_contacts:
        result["E-Mails"] = ir.getEmails()
    if args.number or extract_contacts:
        result["Numbers"] = ir.getPhoneNumber()
    if want_sm or extract_contacts:
        result["SocialMedia"] = ir.getSocials()

    return result


def print_result(result: dict, args: Namespace) -> None:
    print("*" * 50)
    print(f"Target: {result['Target']}")
    print("*" * 50 + "\n")

    if "E-Mails" in result:
        emails = result["E-Mails"]
        print("E-Mails:\n" + ("\n - ".join(emails) if emails else "  (none)"))

    if "Numbers" in result:
        numbers = result["Numbers"]
        print("Numbers:\n" + ("\n - ".join(numbers) if numbers else "  (none)"))

    if "SocialMedia" in result:
        sm = result["SocialMedia"]
        if args.social_extract:
            ir = InfoReader(content=Scrapper(url=result["Target"], crawl=False).getText())
            print("SocialMedia:")
            for entry in ir.getSocialsInfo():
                sm_url, info = entry["url"], entry["info"]
                if info:
                    print(f" - {sm_url}:")
                    for k, v in info.items():
                        print(f"     - {k}: {v}")
                else:
                    print(f" - {sm_url}")
        else:
            print("SocialMedia: " + (", ".join(sm) if sm else "(none)"))

    print()


def save_output(data: dict | list, name: str) -> None:
    Path("output").mkdir(exist_ok=True)
    file_name = name.lower().replace("http://", "").replace("https://", "").replace("/", "")
    out_path = Path("output") / f"{file_name}.json"
    out_path.write_text(json.dumps(data, indent=4))
    print(f"Saved -> {out_path}")


def save_csv_output(results: list, output_path: str) -> None:
    """Save scrape results to a CSV file."""
    Path("output").mkdir(exist_ok=True)
    out_path = Path("output") / output_path

    fieldnames = ["URL", "Emails", "Phone Numbers", "Social Media"]
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for result in results:
            writer.writerow({
                "URL": result.get("Target", ""),
                "Emails": "; ".join(result.get("E-Mails", [])),
                "Phone Numbers": "; ".join(result.get("Numbers", [])),
                "Social Media": "; ".join(result.get("SocialMedia", [])),
            })
    print(f"Saved -> {out_path}")


class IncrementalCsvWriter:
    """Writes results to CSV incrementally so progress isn't lost on crash."""

    def __init__(self, output_path: str):
        Path("output").mkdir(exist_ok=True)
        self.out_path = Path("output") / output_path
        self._fieldnames = ["URL", "Emails", "Phone Numbers", "Social Media"]
        self._file = open(self.out_path, "w", newline="", encoding="utf-8")
        self._writer = csv.DictWriter(self._file, fieldnames=self._fieldnames)
        self._writer.writeheader()
        self._lock = threading.Lock()

    def write(self, result: dict) -> None:
        with self._lock:
            self._writer.writerow({
                "URL": result.get("Target", ""),
                "Emails": "; ".join(result.get("E-Mails", [])),
                "Phone Numbers": "; ".join(result.get("Numbers", [])),
                "Social Media": "; ".join(result.get("SocialMedia", [])),
            })
            self._file.flush()

    def close(self) -> None:
        self._file.close()
        print(f"Saved -> {self.out_path}")


def read_csv_urls(file_path: str, column: str) -> list:
    """Read URLs from a CSV file column."""
    urls = []
    with open(file_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        if column not in reader.fieldnames:
            raise SystemExit(
                f"Column '{column}' not found in CSV. "
                f"Available columns: {', '.join(reader.fieldnames)}"
            )
        for row in reader:
            url = row[column].strip()
            if url:
                urls.append(url)
    return urls


def read_excel_urls(file_path: str, column: str) -> list:
    """Read URLs from an Excel (.xlsx) file column."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise SystemExit("openpyxl is required for Excel support. Install it with: pip install openpyxl")

    wb = load_workbook(file_path, read_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

    if column not in headers:
        raise SystemExit(
            f"Column '{column}' not found in Excel file. "
            f"Available columns: {', '.join(str(h) for h in headers if h)}"
        )

    col_idx = headers.index(column)
    urls = []
    for row in ws.iter_rows(min_row=2):
        val = row[col_idx].value
        if val and str(val).strip():
            urls.append(str(val).strip())
    wb.close()
    return urls


def save_excel_output(results: list, output_path: str) -> None:
    """Save scrape results to an Excel (.xlsx) file."""
    try:
        from openpyxl import Workbook
    except ImportError:
        raise SystemExit("openpyxl is required for Excel support. Install it with: pip install openpyxl")

    Path("output").mkdir(exist_ok=True)
    out_path = Path("output") / output_path

    wb = Workbook()
    ws = wb.active
    ws.title = "Scrape Results"
    ws.append(["URL", "Emails", "Phone Numbers", "Social Media"])

    for result in results:
        ws.append([
            result.get("Target", ""),
            "; ".join(result.get("E-Mails", [])),
            "; ".join(result.get("Numbers", [])),
            "; ".join(result.get("SocialMedia", [])),
        ])

    wb.save(out_path)
    print(f"Saved -> {out_path}")


def scrape_batch(urls: list, args: Namespace, verbose, csv_writer=None) -> list:
    """Scrape multiple URLs concurrently using threads."""
    results = {}
    total = len(urls)
    completed = 0
    failed = 0
    lock = threading.Lock()

    def _scrape_one(url):
        try:
            return url, scrape(url, args)
        except requests.exceptions.RequestException:
            return url, None

    with ThreadPoolExecutor(max_workers=args.threads) as pool:
        futures = {pool.submit(_scrape_one, url): url for url in urls}
        for future in as_completed(futures):
            url, result = future.result()
            with lock:
                completed += 1
                if result is not None:
                    results[url] = result
                    if csv_writer:
                        csv_writer.write(result)
                else:
                    failed += 1
                sys.stderr.write(
                    f"\r[{completed}/{total}] Done: {completed - failed} | Failed: {failed}  "
                )
                sys.stderr.flush()

    sys.stderr.write("\n")
    # Preserve original URL order in output
    return [results[url] for url in urls if url in results]


def main() -> None:
    args = build_parser().parse_args()

    if not args.url and not args.urls and not args.csv:
        raise SystemExit("Please add --url, --urls, or --csv")

    if not args.banner:
        print(BANNER)

    verbose = print if args.verbose else lambda *_: None

    if args.csv:
        csv_path = Path(args.csv)
        if not csv_path.exists():
            raise SystemExit(f"File not found: {args.csv}")

        is_excel = csv_path.suffix.lower() in (".xlsx", ".xls")

        if is_excel:
            raw_urls = read_excel_urls(args.csv, args.csv_column)
        else:
            raw_urls = read_csv_urls(args.csv, args.csv_column)

        urls = [normalize_url(raw) for raw in raw_urls]
        verbose(f"Scraping {len(urls)} URLs with {args.threads} threads...")

        if is_excel:
            results = scrape_batch(urls, args, verbose)
            for result in results:
                print_result(result, args)
            save_excel_output(results, csv_path.stem + "_results.xlsx")
        else:
            writer = IncrementalCsvWriter(csv_path.stem + "_results.csv")
            results = scrape_batch(urls, args, verbose, csv_writer=writer)
            writer.close()
            for result in results:
                print_result(result, args)

    elif args.url:
        url = normalize_url(args.url)
        verbose("Scraping started")
        try:
            result = scrape(url, args)
        except requests.exceptions.RequestException as e:
            raise SystemExit(f"[!] Failed to scrape {url}: {e}")
        verbose("Done")
        print_result(result, args)
        if args.output:
            save_output(result, url)

    else:
        raw_urls = [
            raw.strip() for raw in Path(args.urls).read_text().splitlines()
            if raw.strip()
        ]
        urls = [normalize_url(raw) for raw in raw_urls]
        verbose(f"Scraping {len(urls)} URLs with {args.threads} threads...")
        results = scrape_batch(urls, args, verbose)

        for result in results:
            print_result(result, args)

        if args.output:
            save_output(results, args.urls.replace("/", "_"))


if __name__ == "__main__":
    main()